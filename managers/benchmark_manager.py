import pandas as pd
from sklearn.metrics import accuracy_score, f1_score, precision_score, recall_score, confusion_matrix
import matplotlib.pyplot as plt
from pathlib import Path
from typing import List
import json 
from openpyxl import Workbook
from openpyxl.styles import Font
from typing import List, Dict
from containers.data_container import ModelResponse, ResponseMetadata, DomainAnalysis
from configs.config_loader import get_validated_data, get_output_data_directory, get_model_name_mapping

# TODO 
# Split classes into separate functions/classes based on responsibility, make more modular, and make manager that calls them and handles state.
# Which again can just be called from run_model_benchmarks (main file), like what is done for run_model_interactions. 
# Check if model name and statistic etc is actually present and if not raise error 
# Use mode (single or batch) and batch_size to write output data. Should make a new dir per run, with it's own metadata (timestamp, settings used etc)
# This also applies to the API calls and saving of responses, but ideally we use database instead of writing csv files. 
# Big domains dataset: 
# partition the 900k domains dataset and select around 10k domains to test on (80% malicious & 20% benign): 
# big_dataset = load_dataset("FredZhang7/malicious-website-features-2.4M") 

class ResponseLoader:
    """
    Class to load model response metadata and read the corresponding CSV data, so API calls are uncoupled from analysis.
    This could be extended to read in from DB instead. 
    """
    def __init__(self, mode: str):
        self.metadata_dir = get_output_data_directory() / f"{mode}_processing" / "metadata"

    def load_metadata(self) -> List[ModelResponse]:
        """
        Load metadata files from the metadata directory.

        Returns:
            List[ModelResponse]: A list of ModelResponse instances.
        """
        metadata_files = list(self.metadata_dir.rglob("*.json"))
        model_responses = []
        for metadata_file in metadata_files:
            with open(metadata_file, "r") as f:
                metadata = json.load(f)
                output_data_path = Path(metadata["output_data_path"])
                response_metadata = ResponseMetadata(**metadata["metadata"])
                model_responses.append(ModelResponse(
                    model_name=metadata["model_name"],
                    analysis_results=[],  # don't need full data, keep as empty list
                    metadata=response_metadata,
                    output_data_path=output_data_path
                ))
        print("Loaded metadata:", model_responses)
        return model_responses

    def read_csv_data(self, model_responses: List[ModelResponse]) -> Dict[str, pd.DataFrame]:
        """
        Read the CSV data corresponding to the loaded metadata.

        Args:
            model_responses (List[ModelResponse]): The list of ModelResponse instances.
        Returns:
            Dict[str, pd.DataFrame]: A dictionary of DataFrames keyed by model name.
        """
        model_dfs = {}
        for response in model_responses:
            if response.output_data_path and response.output_data_path.exists():
                df = pd.read_csv(response.output_data_path)
                model_dfs[response.model_name] = df
        return model_dfs

    def create_model_responses(self, model_responses: List[ModelResponse], model_dfs: Dict[str, pd.DataFrame]) -> List[ModelResponse]:
        """
        Create ModelResponse instances from metadata and CSV data.

        Args:
            model_responses (List[ModelResponse]): The list of ModelResponse instances.
            model_dfs (Dict[str, pd.DataFrame]): The dictionary of DataFrames keyed by model name.
        Returns:
            List[ModelResponse]: A list of ModelResponse instances with filled analysis_results.
        """
        for response in model_responses:
            df = model_dfs[response.model_name]
            analysis_results = []
            for _, row in df.iterrows():
                reasons = row["reasons"]
                # handle None or NaN values in reasons
                if pd.isna(reasons):
                    reasons = None
                else:
                    reasons = reasons.split(", ")
                analysis_results.append(
                    DomainAnalysis(
                        domain_name=row["domain_name"],
                        category=row["category"],
                        confidence_score=row["confidence_score"],
                        reasons=reasons
                    )
                )
            response.analysis_results = analysis_results
        print("Created model responses:", model_responses)
        return model_responses


class ResponseManager:
    """
    Class to manage the loading, merging, and analysis of model responses.
    """
    def __init__(self, validated_df: pd.DataFrame):
        self.validated_df = validated_df

    def merge_with_validated_data(self, model_dfs: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
        """
        Merge the model response DataFrames with the validated data.

        Args:
            model_dfs (Dict[str, pd.DataFrame]): The dictionary of model response DataFrames.
        Returns:
            Dict[str, pd.DataFrame]: A dictionary of merged DataFrames.
        """
        merged_dfs = {}
        for model_name, df in model_dfs.items():
            merged_df = df.merge(self.validated_df, left_on="domain_name", right_on="fqdn", suffixes=("_model", "_validated"))
            merged_dfs[model_name] = merged_df
        print("Merged data:", merged_dfs)
        return merged_dfs

    def calculate_statistics(self, merged_dfs: Dict[str, pd.DataFrame]) -> Dict[str, Dict[str, float]]:
        """
        Calculate statistics for the merged DataFrames.

        Args:
            merged_dfs (Dict[str, pd.DataFrame]): The dictionary of merged DataFrames.
        Returns:
            Dict[str, Dict[str, float]]: A dictionary of calculated statistics.
        """
        statistics = {}
        for model_name, df in merged_dfs.items():
            y_true = df["category_validated"]
            y_pred = df["category_model"]
            
            accuracy = accuracy_score(y_true, y_pred)
            f1 = f1_score(y_true, y_pred, pos_label="malicious", average="binary")
            precision = precision_score(y_true, y_pred, pos_label="malicious", average="binary")
            recall = recall_score(y_true, y_pred, pos_label="malicious", average="binary")
            tn, fp, fn, tp = confusion_matrix(y_true, y_pred, labels=["benign", "malicious"]).ravel()

            statistics[model_name] = {
                "accuracy": accuracy,
                "f1_score": f1,
                "precision": precision,
                "recall": recall,
                "false_positives": fp,
                "false_negatives": fn,
                "num_domains": len(df)
            }
        print("Calculated statistics:", statistics)
        return statistics

    def save_numerical_output(self, statistics: Dict[str, Dict[str, float]], output_dir: Path):
        """
        Save the calculated statistics to a CSV file.

        Args:
            statistics (Dict[str, Dict[str, float]]): The dictionary of calculated statistics.
            output_dir (Path): The directory where the numerical output will be saved.
        """
        output_dir.mkdir(parents=True, exist_ok=True)
        
        for model_name, stats in statistics.items():
            df = pd.DataFrame([stats])
            file_path = output_dir / f"{model_name}_statistics.csv"
            df.to_csv(file_path, index=False)
            print(f"Numerical output saved to {file_path}")

    def plot_statistic(self, statistics: Dict[str, Dict[str, float]], stat_name: str, title: str, save_plots: bool, data_dir: Path):
        """
        Plot a specified statistic for the models.

        Args:
            statistics (Dict[str, Dict[str, float]]): The dictionary of calculated statistics.
            stat_name (str): The name of the statistic to plot.
            title (str): The title of the plot.
            save_plots (bool): Whether to save the plots or just show them.
            data_dir (Path): The directory where the data CSV files are stored.
        """
        model_names = list(statistics.keys())
        stat_values = [stats[stat_name] for stats in statistics.values()]
        num_domains = [stats['num_domains'] for stats in statistics.values()]
        
        plt.figure(figsize=(10, 6))
        bars = plt.bar(model_names, stat_values)
        plt.title(title)
        plt.xlabel("Model")
        plt.ylabel(stat_name.replace('_', ' ').title())
        plt.xticks(rotation=90)
        
        # Add text above bars
        for bar, stat_value, num_domain in zip(bars, stat_values, num_domains):
            yval = bar.get_height()
            plt.text(bar.get_x() + bar.get_width() / 2, yval, f"{stat_value:.2f}\n({num_domain})", ha="center", va="bottom")
        
        plt.tight_layout()
        if save_plots:
            output_dir = data_dir.parent / "plots"
            output_dir.mkdir(parents=True, exist_ok=True)
            plot_path = output_dir / f"{title.replace(' ', '_').lower()}.png"
            plt.savefig(plot_path)
            print(f"Plot saved to {plot_path}")
        else:
            plt.show()

    def plot_all_statistics(self, statistics: Dict[str, Dict[str, float]], save_plots: bool, data_dir: Path):
        """
        Plot all statistics for the models.

        Args:
            statistics (Dict[str, Dict[str, float]]): The dictionary of calculated statistics.
            save_plots (bool): Whether to save the plots or just show them.
            data_dir (Path): The directory where the data CSV files are stored.
        """
        self.plot_statistic(statistics, "accuracy", "Accuracy", save_plots, data_dir)
        self.plot_statistic(statistics, "f1_score", "F1 Score", save_plots, data_dir)
        self.plot_statistic(statistics, "precision", "Precision", save_plots, data_dir)
        self.plot_statistic(statistics, "recall", "Recall", save_plots, data_dir)
        self.plot_statistic(statistics, "false_positives", "False Positives", save_plots, data_dir)
        self.plot_statistic(statistics, "false_negatives", "False Negatives", save_plots, data_dir)

class ModelDataframeComparison:
    """
    Class to compare model categorizations with validated data and highlight the results.
    """
    def __init__(self, merged_dfs: Dict[str, pd.DataFrame], selected_models: List[str] = None):
        self.merged_dfs = merged_dfs
        self.selected_models = selected_models

    def highlight_correct_incorrect(self) -> pd.DataFrame:
        """
        Highlight the model categorizations based on correctness and combine them into a single DataFrame.

        Returns:
            pd.DataFrame: A unified DataFrame with categorizations and confidence scores for all models.
        """
        unified_df = None
        
        for model_name, df in self.merged_dfs.items():
            if self.selected_models and model_name not in self.selected_models:
                continue
            
            df = df.copy()
            df[f"{model_name}_category_model"] = df["category_model"]
            df[f"{model_name}_confidence_score"] = df["confidence_score"]
            
            if unified_df is None:
                unified_df = df[["domain_name", "category_validated", f"{model_name}_category_model", f"{model_name}_confidence_score"]]
            else:
                unified_df = unified_df.merge(
                    df[["domain_name", f"{model_name}_category_model", f"{model_name}_confidence_score"]],
                    on="domain_name", how="outer"
                )
        
        return unified_df

    def save_highlighted_to_excel(self, output_dir: Path):
        """
        Save the unified highlighted DataFrame to an Excel file with colored cells in the specified directory.

        Args:
            output_dir (Path): The directory to save the Excel file.
        """
        unified_df = self.highlight_correct_incorrect()
        output_dir.mkdir(parents=True, exist_ok=True)
        
        excel_path = output_dir / "combined_highlighted.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Highlighted Results"  # type: ignore
        
        # write header
        headers = list(unified_df.columns)
        sheet.append(headers)  # type: ignore
        
        # write data with highlighted categories
        for _, row in unified_df.iterrows():
            row_data = []
            for col in unified_df.columns:
                row_data.append(row[col])
                
            sheet.append(row_data)  # type: ignore
            
            # Apply font color to the category_model cells
            for col_idx in range(2, len(headers), 2):  # start from 3rd column and step by 2
                cell = sheet.cell(row=sheet.max_row, column=col_idx + 1)  # 1-based indexing for columns
                category_model = row[headers[col_idx]]
                category_validated = row["category_validated"]
                
                if category_model == category_validated:
                    cell.font = Font(color="00FF00")
                else:
                    cell.font = Font(color="FF0000")
        
        workbook.save(excel_path)
        print(f"Combined highlighted Excel saved to {excel_path}")


def run_highlighted_comparison(selected_models: List[str] = ["all"], mode: str = "single"):
    """
    Run the highlighted comparison for the models.

    Args:
        selected_models (List[str]): List of model names to analyze. Use ["all"] for all models.
        mode (str): Mode of processing, either "single" or "batch".
        config_path (str): Path to the JSON configuration file.
    """
    # Load the model name mapping from the JSON configuration
    MODEL_NAME_MAPPING = get_model_name_mapping()
    
    # Initialize response loader with mode
    response_loader = ResponseLoader(mode)
    
    # Load metadata and read CSV data
    model_responses = response_loader.load_metadata()
    print("Model responses loaded:", model_responses)
    
    # Filter the models if not "all" is selected
    if selected_models != ["all"]:
        # Convert shortened model names to full names using the mapping
        selected_models_full_names = [MODEL_NAME_MAPPING.get(name, name) for name in selected_models]
        model_responses = [md for md in model_responses if md.model_name in selected_models_full_names]
    print("Filtered model responses:", model_responses)
    
    if not model_responses:
        print("No models selected after filtering.")
        return
    
    # Read CSV (response) data for the filtered models
    model_dfs = response_loader.read_csv_data(model_responses)
    print("Model dataframes:", model_dfs)
    
    # Create ModelResponse instances with the loaded data
    model_responses = response_loader.create_model_responses(model_responses, model_dfs)
    print("Created model responses:", model_responses)
    
    # Load the validated data
    validated_df = pd.read_csv(get_validated_data())
    print("Validated data loaded:", validated_df.head())
    
    # Initialize the ResponseManager with the validated DataFrame
    response_manager = ResponseManager(validated_df)
    
    # Merge with validated data
    merged_dfs = response_manager.merge_with_validated_data(model_dfs)
    print("Merged dataframes:", merged_dfs)
    
    if not merged_dfs:
        print("Merging failed, no data available.")
        return
    
    # Highlighted comparison
    model_comparison = ModelDataframeComparison(merged_dfs, selected_models_full_names if selected_models != ["all"] else None)
    output_highlighted_dir = get_output_data_directory() / f"{mode}_processing" / "highlighted_comparisons"
    model_comparison.save_highlighted_to_excel(output_highlighted_dir)


def run_analysis(selected_models: List[str] = ["all"], mode: str = "single", analyses: List[str] = ["all"], save_plots: bool = True, numerical_output: bool = False):
    """
    Run the analysis and plotting pipeline.

    Args:
        selected_models (List[str]): List of model names to analyze. Use ["all"] for all models.
        mode (str): Mode of processing, either "single" or "batch".
        analyses (List[str]): List of analyses to run. Use ["all"] for all analyses. Available analyses: 
        save_plots (bool): Whether to save the plots or just show them.
        numerical_output (bool): Whether to save the numerical output to files.
    """
    # Load the model name mapping from the JSON configuration
    MODEL_NAME_MAPPING = get_model_name_mapping()
    
    # Initialize the ResponseLoader with mode
    response_loader = ResponseLoader(mode)
    
    # Load metadata
    model_responses = response_loader.load_metadata()
    print("Model responses loaded:", model_responses)
    
    # Filter the models if not "all" is selected
    if selected_models != ["all"]:
        # convert short model names to full names using the mapping
        selected_models_full_names = [MODEL_NAME_MAPPING.get(name, name) for name in selected_models]
        model_responses = [md for md in model_responses if md.model_name in selected_models_full_names]
    print("Filtered model responses:", model_responses)
    
    if not model_responses:
        print("No models selected after filtering.")
        return
    
    # Read CSV data for the filtered models
    model_dfs = response_loader.read_csv_data(model_responses)
    print("Model dataframes:", model_dfs)
    
    # Create ModelResponse instances with the loaded data
    model_responses = response_loader.create_model_responses(model_responses, model_dfs)
    print("Created model responses:", model_responses)
    
    # Load the validated data
    validated_df = pd.read_csv(get_validated_data())
    print("Validated data loaded:", validated_df.head())
    
    # Initialize the ResponseManager with the validated DataFrame
    response_manager = ResponseManager(validated_df)
    
    # Merge with validated data
    merged_dfs = response_manager.merge_with_validated_data(model_dfs)
    print("Merged dataframes:", merged_dfs)
    
    if not merged_dfs:
        print("Merging failed, no data available.")
        return
    
    # Calculate statistics
    statistics = response_manager.calculate_statistics(merged_dfs)
    print("Calculated statistics:", statistics)
    
    if not statistics:
        print("Statistics calculation failed, no data available.")
        return
    
    # Get the directory where the data CSV files are stored
    data_dir = get_output_data_directory() / f"{mode}_processing" / "model_responses_csv"
    
    # Save numerical output if enabled
    if numerical_output:
        numerical_output_dir = get_output_data_directory() / f"{mode}_processing" / "numerical_output"
        response_manager.save_numerical_output(statistics, numerical_output_dir)
    
    # Run plotting based on selected analyses
    if analyses == ["all"]:
        response_manager.plot_all_statistics(statistics, save_plots, data_dir)
    else:
        for analysis in analyses:
            if analysis in ["accuracy", "f1_score", "precision", "recall", "false_positives", "false_negatives"]:
                response_manager.plot_statistic(statistics, analysis, analysis.replace('_', ' ').title(), save_plots, data_dir)
            else:
                print(f"Invalid analysis selected: {analysis}")

